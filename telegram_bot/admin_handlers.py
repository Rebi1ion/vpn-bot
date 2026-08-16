from aiogram import Router, types, F, Bot
from aiogram.filters.command import Command
from aiogram.fsm.context import FSMContext
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from sqlalchemy import delete, func, select
from database.models import Channel, User, UserChannelSubscription, VpnConfig
from database.server_operations import (
    add_server,
    delete_server,
    get_all_servers,
    get_server_by_id,
    sync_server_count,
    update_server,
)
from telegram_bot.keyboards import (
    get_admin_panel_keyboard,
    get_main_keyboard,
    get_admin_keyboard,
    get_cancel_keyboard,
    get_channel_actions_keyboard,
    get_server_actions_keyboard,
    get_server_management_keyboard,
    get_settings_keyboard,
    get_user_manage_keyboard,
)
from telegram_bot.states import (
    AdminStates,
    ServerStates,
    UserManagementStates,
    AdminSettingsStates,
)
from database import AsyncSessionLocal
from database.operations import (
    create_channel,
    get_all_active_channels,
    delete_channel,
    get_channel_by_username,
    count_active_channels,
    get_user_by_telegram_id,
    get_user_by_username,
    get_user_config,
    get_setting,
    set_setting,
)
from config.settings import settings
import logging

logger = logging.getLogger(__name__)
router = Router()


def is_admin(message: types.Message) -> bool:
    """Фильтр для проверки что пользователь админ"""
    return message.from_user.id == settings.ADMIN_USER_ID


# ============ АДМИН ПАНЕЛЬ ============


@router.message(F.text == "⚙️ Админ панель")
async def admin_panel(message: types.Message):
    """Открыть админ панель"""
    if not is_admin(message):
        await message.answer("❌ У вас нет доступа к админ панели")
        return

    async with AsyncSessionLocal() as session:
        channels_count = await count_active_channels(session)
        price_1m = await get_setting(session, "price_1_month_rub", 150)
        price_6m = await get_setting(session, "price_6_months_rub", 700)

    await message.answer(
        f"⚙️ <b>Админ панель</b>\n\n"
        f"📋 Всего каналов: <b>{channels_count}</b>\n\n"
        f"💰 Текущие цены:\n"
        f"🔹 1 месяц: <b>{price_1m} руб.</b>\n"
        f"🔹 6 месяцев: <b>{price_6m} руб.</b>",
        reply_markup=get_admin_panel_keyboard(),
        parse_mode="HTML",
    )


@router.message(F.text == "🔙 Назад")
async def back_to_main(message: types.Message):
    """Вернуться в главное меню"""
    if not is_admin(message):
        return

    # Проверяем админ или нет
    if message.from_user.id == settings.ADMIN_USER_ID:
        keyboard = get_admin_keyboard()
    else:
        keyboard = get_main_keyboard()

    await message.answer("Главное меню", reply_markup=keyboard)


# ============ ДОБАВЛЕНИЕ КАНАЛА ============


@router.message(F.text == "➕ Добавить канал")
async def add_channel_start(message: types.Message, state: FSMContext):
    """Начать процесс добавления канала"""
    if not is_admin(message):
        return

    await message.answer(
        "📝 Отправьте канал одним из способов:\n\n"
        "📢 Публичный канал:\n"
        "- https://t.me/channel_name\n\n"
        "🔒 Приватный канал:\n"
        "- https://t.me/+XXXXX (invite ссылка)\n"
        "- -1001234567890 (Chat ID)",
        reply_markup=get_cancel_keyboard(),
    )
    await state.set_state(AdminStates.waiting_for_channel_username)


@router.message(AdminStates.waiting_for_channel_username)
async def add_channel_username(message: types.Message, state: FSMContext):
    """Получили Chat ID или username канала"""
    if message.text == "❌ Отмена":
        await state.clear()
        await message.answer("❌ Отменено", reply_markup=get_admin_panel_keyboard())
        return

    input_text = message.text.strip()

    # Проверяем формат
    if input_text.startswith("@"):
        # Публичный канал
        channel_username = input_text
        channel_id = input_text
        is_private = False
    elif input_text.startswith("-"):
        # Приватный канал (Chat ID)
        if not input_text[1:].isdigit():
            await message.answer(
                "❌ <b>Неправильный формат Chat ID!</b>\n\n"
                "Chat ID должен:\n"
                "- Начинаться с <code>-</code>\n"
                "- Содержать только цифры\n\n"
                "Пример: <code>-1001234567890</code>\n\n"
                "Попробуй ещё раз или нажми ❌ Отмена",
                parse_mode="HTML",
            )
            return

        channel_id = input_text
        channel_username = input_text  # Для приватных каналов username = chat_id
        is_private = True
    else:
        await message.answer(
            "❌ <b>Неправильный формат!</b>\n\n"
            "Введи:\n"
            "- <code>@username</code> для публичного канала\n"
            "- <code>-1001234567890</code> для приватного\n\n"
            "Попробуй ещё раз или нажми ❌ Отмена",
            parse_mode="HTML",
        )
        return

    # Сохраняем данные
    await state.update_data(
        channel_username=channel_username, channel_id=channel_id, is_private=is_private
    )

    await state.set_state(AdminStates.waiting_for_invite_link)

    if is_private:
        await message.answer(
            "✅ <b>Chat ID сохранён!</b>\n\n"
            "<b>Шаг 2 из 3:</b> Отправь invite-ссылку на канал\n\n"
            "Пример: <code>https://t.me/+AbCdEfGhIjKlMnO</code>",
            parse_mode="HTML",
        )
    else:
        await message.answer(
            "✅ <b>Username сохранён!</b>\n\n"
            "<b>Шаг 2 из 3:</b> Отправь ссылку на канал\n\n"
            "Можешь отправить:\n"
            "- Username: <code>@username</code>\n"
            "- Или URL: <code>https://t.me/username</code>",
            parse_mode="HTML",
        )


@router.message(AdminStates.waiting_for_invite_link)
async def add_invite_link(message: types.Message, state: FSMContext):
    """Получили invite ссылку для приватного канала с Chat ID"""
    if message.text == "❌ Отмена":
        await state.clear()
        await message.answer("❌ Отменено", reply_markup=get_admin_panel_keyboard())
        return

    invite_link = message.text.strip()

    # Проверяем что это ссылка
    if not ("t.me/" in invite_link or "telegram.me/" in invite_link):
        await message.answer(
            "❌ Это не похоже на ссылку!\n"
            "Отправь invite ссылку вида:\n"
            "https://t.me/+XXXXX",
            reply_markup=get_cancel_keyboard(),
        )
        return

    # Сохраняем invite ссылку для кнопки
    await state.update_data(channel_url=invite_link)

    data = await state.get_data()
    channel_title = data.get("channel_title", "Канал")

    await message.answer(
        f"✅ Invite ссылка добавлена!\n\n"
        f"📋 Канал: {channel_title}\n"
        f"🔗 Ссылка: {invite_link}\n\n"
        f"📝 Теперь отправь название канала\n"
        f"(то что будет показано пользователям на кнопке)",
        reply_markup=get_cancel_keyboard(),
    )
    await state.set_state(AdminStates.waiting_for_channel_name)


@router.message(AdminStates.waiting_for_channel_name)
async def add_channel_name(message: types.Message, state: FSMContext):
    """Получили название канала"""
    if message.text == "❌ Отмена":
        await state.clear()
        await message.answer("❌ Отменено", reply_markup=get_admin_panel_keyboard())
        return

    display_name = message.text.strip()
    data = await state.get_data()

    channel_username = data.get("channel_username")
    channel_url = data.get("channel_url", channel_username)

    async with AsyncSessionLocal() as session:
        # Проверяем что канал не существует
        existing = await get_channel_by_username(session, channel_username)

        if existing:
            # Канал уже существует
            await message.answer(
                "❌ <b>Этот канал уже добавлен!</b>\n\n"
                f"<b>Название:</b> {existing.display_name}\n"
                f"<b>Username:</b> <code>{existing.channel_username}</code>\n\n"
                f"Если хочешь изменить канал, сначала удали старый.",
                parse_mode="HTML",
                reply_markup=get_admin_panel_keyboard(),
            )
            await state.clear()
            return

        # Создаём новый канал в БД
        channel = await create_channel(
            session=session,
            channel_username=channel_username,
            channel_url=channel_url,
            display_name=display_name,
            order=0,
        )

        await state.clear()

        await message.answer(
            f"✅ <b>Канал успешно добавлен!</b>\n\n"
            f"<b>Название:</b> {channel.display_name}\n"
            f"<b>Username:</b> <code>{channel.channel_username or 'Не указан'}</code>\n"
            f"<b>URL:</b> {channel.channel_url}",
            parse_mode="HTML",
            reply_markup=get_admin_panel_keyboard(),
        )


# ============ СПИСОК КАНАЛОВ ============


@router.message(F.text == "📋 Список каналов")
async def list_channels(message: types.Message):
    """Показать список всех каналов"""
    if not is_admin(message):
        return

    async with AsyncSessionLocal() as session:
        channels = await get_all_active_channels(session)

    if not channels:
        await message.answer(
            "📋 Каналов пока нет\n\n" "Добавьте первый канал через '➕ Добавить канал'",
            reply_markup=get_admin_panel_keyboard(),
        )
        return

    await message.answer(
        f"📋 Всего каналов: {len(channels)}\n"
        f"Очерёдность выдачи — по времени добавления"
    )

    for i, channel in enumerate(channels, 1):
        channel_type = (
            "🔒 Приватный"
            if channel.channel_username.startswith("-100")
            else "📢 Публичный"
        )

        await message.answer(
            f"{i}. {channel.display_name} {channel_type}\n"
            f"🆔 ID: {channel.channel_username}\n"
            f"🔗 Ссылка: {channel.channel_url}\n"
            f"⭐️ Партнёрский: {'✅ Да' if channel.is_partner_channel else '❌ Нет'}\n"
            f"🕐 Добавлен: {channel.created_at.strftime('%d.%m.%Y %H:%M')}",
            reply_markup=get_channel_actions_keyboard(
                channel.id, channel.is_partner_channel
            ),
        )


# ============ УДАЛЕНИЕ КАНАЛА ============


@router.callback_query(F.data.startswith("delete_channel:"))
async def delete_channel_callback(callback: types.CallbackQuery):
    """Обработка удаления канала"""
    if callback.from_user.id != settings.ADMIN_USER_ID:
        await callback.answer("❌ Нет доступа", show_alert=True)
        return

    channel_id = int(callback.data.split(":")[1])

    async with AsyncSessionLocal() as session:
        success = await delete_channel(session, channel_id)

    if success:
        await callback.message.delete()
        await callback.answer("✅ Канал удалён")
        logger.info(f"🗑️ Админ удалил канал ID: {channel_id}")
    else:
        await callback.answer("❌ Ошибка удаления", show_alert=True)


@router.callback_query(F.data.startswith("toggle_partner_channel:"))
async def toggle_partner_channel_callback(callback: types.CallbackQuery):
    """Сделать канал партнерским или наоборот"""
    if callback.from_user.id != settings.ADMIN_USER_ID:
        await callback.answer("❌ Нет доступа", show_alert=True)
        return

    channel_id = int(callback.data.split(":")[1])

    async with AsyncSessionLocal() as session:
        channel = await session.get(Channel, channel_id)
        if channel:
            # Опционально: можно снимать флаг с остальных, если партнерский канал может быть только один
            channel.is_partner_channel = not channel.is_partner_channel
            await session.commit()

            # Обновляем клавиатуру
            await callback.message.edit_reply_markup(
                reply_markup=get_channel_actions_keyboard(
                    channel.id, channel.is_partner_channel
                )
            )

            status = "партнёрским" if channel.is_partner_channel else "обычным"
            await callback.answer(f"✅ Канал теперь {status}")

            # Обновляем текст сообщения, если возможно
            lines = callback.message.text.split("\n")
            for i, line in enumerate(lines):
                if line.startswith("⭐️ Партнёрский:"):
                    lines[i] = (
                        f"⭐️ Партнёрский: {'✅ Да' if channel.is_partner_channel else '❌ Нет'}"
                    )
                    break

            await callback.message.edit_text(
                "\n".join(lines),
                reply_markup=get_channel_actions_keyboard(
                    channel.id, channel.is_partner_channel
                ),
            )


# ============ СТАТИСТИКА ============


@router.message(F.text == "📊 Статистика")
async def show_stats(message: types.Message):
    """Показать статистику"""
    if not is_admin(message):
        return

    async with AsyncSessionLocal() as session:
        from sqlalchemy import func, select
        from database.models import User, VpnConfig

        total_users = await session.execute(select(func.count(User.id)))
        total_users = total_users.scalar()

        active_configs = await session.execute(
            select(func.count(VpnConfig.id)).where(VpnConfig.is_active == True)
        )
        active_configs = active_configs.scalar()

        inactive_configs = await session.execute(
            select(func.count(VpnConfig.id)).where(VpnConfig.is_active == False)
        )
        inactive_configs = inactive_configs.scalar()

        channels_count = await count_active_channels(session)

    await message.answer(
        f"📊 Статистика\n\n"
        f"👥 Всего пользователей: {total_users}\n"
        f"✅ Активных конфигов: {active_configs}\n"
        f"❌ Отключённых конфигов: {inactive_configs}\n"
        f"📋 Каналов: {channels_count}",
        reply_markup=get_admin_panel_keyboard(),
    )


# ============ ПОЛУЧЕНИЕ CHAT ID ============


@router.message(Command("get_channel_id"))
async def get_channel_id_command(message: types.Message):
    """Получить Chat ID из пересланного сообщения"""
    if not is_admin(message):
        return

    # Проверяем что это переслано из канала
    if not message.forward_from_chat:
        await message.answer(
            "❌ Команда работает только с пересланными сообщениями!\n\n"
            "Инструкция:\n"
            "1. Добавь бота админом в приватный канал\n"
            "2. Перешли сюда любое сообщение из этого канала\n"
            "3. Вместе с сообщением напиши /get_channel_id"
        )
        return

    chat = message.forward_from_chat
    chat_id = chat.id
    chat_title = chat.title or "Неизвестный"
    chat_type = "Канал" if chat.type == "channel" else "Группа"

    await message.answer(
        f"✅ Получен Chat ID!\n\n"
        f"📋 Название: {chat_title}\n"
        f"📁 Тип: {chat_type}\n"
        f"🆔 Chat ID: <code>{chat_id}</code>\n\n"
        f"💡 Скопируй этот ID и используй при добавлении канала",
        parse_mode="HTML",
    )


@router.message(Command("stats"))
async def show_detailed_stats(message: types.Message):
    """Подробная статистика для админа"""
    if not is_admin(message):
        return

    try:
        async with AsyncSessionLocal() as session:
            from sqlalchemy import func
            from datetime import datetime, timedelta
            import psutil
            import os

            # Статистика пользователей
            total_users = await session.scalar(select(func.count(User.id)))

            # Новые пользователи за 24ч
            yesterday = datetime.utcnow() - timedelta(days=1)
            new_users_24h = await session.scalar(
                select(func.count(User.id)).where(User.created_at >= yesterday)
            )

            # Конфиги
            active_configs = await session.scalar(
                select(func.count(VpnConfig.id)).where(VpnConfig.is_active == True)
            )

            inactive_configs = await session.scalar(
                select(func.count(VpnConfig.id)).where(VpnConfig.is_active == False)
            )

            # Каналы
            total_channels = await session.scalar(
                select(func.count(Channel.id)).where(Channel.is_active == True)
            )

            # Системные ресурсы
            process = psutil.Process()
            mem_info = process.memory_info()
            mem_mb = mem_info.rss / 1024 / 1024

            cpu_percent = process.cpu_percent(interval=0.5)

            # Диск
            disk = psutil.disk_usage(".")
            disk_free_gb = disk.free / 1024 / 1024 / 1024
            disk_percent = disk.percent

            # Размер БД
            db_size = os.path.getsize("vpn_bot.db") / 1024 / 1024

            # Бэкапы
            from backup import backup_manager

            backup_info = backup_manager.get_backup_info()

            # Uptime
            uptime_seconds = time.time() - process.create_time()
            uptime_hours = uptime_seconds / 3600

        text = (
            "📊 <b>Детальная статистика</b>\n\n"
            "<b>👥 Пользователи:</b>\n"
            f"- Всего: {total_users}\n"
            f"- Новых за 24ч: {new_users_24h}\n\n"
            "<b>🔐 VPN конфиги:</b>\n"
            f"- Активных: {active_configs}\n"
            f"- Отключённых: {inactive_configs}\n\n"
            "<b>📋 Каналы:</b>\n"
            f"- Всего: {total_channels}\n\n"
            "<b>💻 Система:</b>\n"
            f"- RAM: {mem_mb:.1f} MB\n"
            f"- CPU: {cpu_percent:.1f}%\n"
            f"- Диск: {disk_free_gb:.1f} GB свободно ({disk_percent}%)\n"
            f"- БД: {db_size:.2f} MB\n"
            f"- Uptime: {uptime_hours:.1f}ч\n\n"
            "<b>💾 Бэкапы:</b>\n"
            f"- Количество: {backup_info['count']}\n"
            f"- Размер: {backup_info['total_size_mb']:.1f} MB\n"
            f"- Последний: {backup_info['latest'] or 'нет'}"
        )

        await message.answer(text, parse_mode="HTML")

    except Exception as e:
        logger.error(f"Ошибка получения статистики: {e}", exc_info=True)
        await message.answer("❌ Ошибка получения статистики")


@router.message(Command("backup"))
async def manual_backup(message: types.Message):
    """Создать бэкап вручную"""
    if not is_admin(message):
        return

    try:
        await message.answer("💾 Создаю резервную копию...")

        from backup import backup_manager

        success = await backup_manager.create_backup()

        if success:
            info = backup_manager.get_backup_info()
            await message.answer(
                f"✅ Бэкап создан!\n\n"
                f"Всего бэкапов: {info['count']}\n"
                f"Последний: {info['latest']}"
            )
        else:
            await message.answer("❌ Не удалось создать бэкап")

    except Exception as e:
        logger.error(f"Ошибка ручного бэкапа: {e}", exc_info=True)
        await message.answer("❌ Ошибка создания бэкапа")


# Импорт для статистики
import time


# ============ УПРАВЛЕНИЕ СЕРВЕРАМИ ============


@router.message(F.text == "🖥️ Управление серверами")
async def server_management_menu(message: types.Message):
    """Меню управления серверами"""
    if not is_admin(message):
        return

    keyboard = get_server_management_keyboard()
    await message.answer(
        "🖥️ <b>Управление серверами</b>\n\n" "Выбери действие:",
        reply_markup=keyboard,
        parse_mode="HTML",
    )


@router.message(F.text == "📋 Список серверов")
async def list_servers(message: types.Message):
    """Показать список серверов"""
    if not is_admin(message):
        return

    async with AsyncSessionLocal() as session:
        servers = await get_all_servers(session)

        if not servers:
            await message.answer("📋 Нет серверов в системе")
            return

        for server in servers:
            status = "🟢 Активен" if server.is_active else "🔴 Отключён"
            load = (
                (server.current_users / server.max_users * 100)
                if server.max_users > 0
                else 0
            )

            if load < 50:
                load_indicator = "🟢"
            elif load < 85:
                load_indicator = "🟡"
            else:
                load_indicator = "🔴"

            text = (
                f"<b>🖥️ {server.name}</b>\n\n"
                f"Статус: {status}\n"
                f"IP: <code>{server.ip}</code>\n"
                f"Host: <code>{server.host}</code>\n"
                f"Inbound ID: {server.inbound_id}\n"
                f"Загрузка: {load_indicator} {server.current_users}/{server.max_users} ({load:.1f}%)\n"
                f"ID: <code>{server.id}</code>"
            )

            keyboard = get_server_actions_keyboard(server.id)
            await message.answer(text, reply_markup=keyboard, parse_mode="HTML")


@router.message(F.text == "➕ Добавить сервер")
async def add_server_start(message: types.Message, state: FSMContext):
    """Начать добавление сервера"""
    if not is_admin(message):
        return

    await state.set_state(ServerStates.waiting_for_name)

    cancel_kb = get_cancel_keyboard()
    await message.answer(
        "📝 <b>Добавление нового сервера</b>\n\n"
        "Введи название сервера (например: Server 1):",
        reply_markup=cancel_kb,
        parse_mode="HTML",
    )


@router.message(ServerStates.waiting_for_name)
async def process_server_name(message: types.Message, state: FSMContext):
    """Обработка названия"""
    if message.text == "❌ Отмена":
        await state.clear()
        keyboard = get_admin_panel_keyboard()
        await message.answer("❌ Добавление сервера отменено", reply_markup=keyboard)
        return

    await state.update_data(name=message.text)
    await state.set_state(ServerStates.waiting_for_ip)

    await message.answer(
        "📝 Введи IP адрес сервера (например: 123.45.67.89):", parse_mode="HTML"
    )


@router.message(ServerStates.waiting_for_ip)
async def process_server_ip(message: types.Message, state: FSMContext):
    """Обработка IP"""
    if message.text == "❌ Отмена":
        await state.clear()
        keyboard = get_admin_panel_keyboard()
        await message.answer("❌ Добавление сервера отменено", reply_markup=keyboard)
        return

    await state.update_data(ip=message.text)
    await state.set_state(ServerStates.waiting_for_host)

    await message.answer(
        "📝 Введи HOST (URL панели 3x-ui, например: https://123.45.67.89:2053):",
        parse_mode="HTML",
    )


@router.message(ServerStates.waiting_for_host)
async def process_server_host(message: types.Message, state: FSMContext):
    """Обработка HOST"""
    if message.text == "❌ Отмена":
        await state.clear()
        keyboard = get_admin_panel_keyboard()
        await message.answer("❌ Добавление сервера отменено", reply_markup=keyboard)
        return

    await state.update_data(host=message.text)
    await state.set_state(ServerStates.waiting_for_username)

    await message.answer("📝 Введи USERNAME (логин панели 3x-ui):", parse_mode="HTML")


@router.message(ServerStates.waiting_for_username)
async def process_server_username(message: types.Message, state: FSMContext):
    """Обработка USERNAME"""
    if message.text == "❌ Отмена":
        await state.clear()
        keyboard = get_admin_panel_keyboard()
        await message.answer("❌ Добавление сервера отменено", reply_markup=keyboard)
        return

    await state.update_data(username=message.text)
    await state.set_state(ServerStates.waiting_for_password)

    await message.answer("📝 Введи PASSWORD (пароль панели 3x-ui):", parse_mode="HTML")


@router.message(ServerStates.waiting_for_password)
async def process_server_password(message: types.Message, state: FSMContext):
    """Обработка PASSWORD"""
    if message.text == "❌ Отмена":
        await state.clear()
        keyboard = get_admin_panel_keyboard()
        await message.answer("❌ Добавление сервера отменено", reply_markup=keyboard)
        return

    await state.update_data(password=message.text)
    await state.set_state(ServerStates.waiting_for_inbound_id)

    await message.answer("📝 Введи INBOUND ID (обычно 1):", parse_mode="HTML")


@router.message(ServerStates.waiting_for_inbound_id)
async def process_server_inbound_id(message: types.Message, state: FSMContext):
    """Обработка INBOUND ID и создание сервера"""
    if message.text == "❌ Отмена":
        await state.clear()
        keyboard = get_admin_panel_keyboard()
        await message.answer("❌ Добавление сервера отменено", reply_markup=keyboard)
        return

    try:
        inbound_id = int(message.text)
    except ValueError:
        await message.answer("❌ INBOUND ID должен быть числом. Попробуй снова:")
        return

    # Получаем все данные
    data = await state.get_data()

    async with AsyncSessionLocal() as session:
        server = await add_server(
            session=session,
            name=data["name"],
            ip=data["ip"],
            host=data["host"],
            username=data["username"],
            password=data["password"],
            inbound_id=inbound_id,
            max_users=350,
            # priority=1
        )

    await state.clear()

    keyboard = get_admin_panel_keyboard()
    await message.answer(
        f"✅ <b>Сервер добавлен!</b>\n\n"
        f"Название: {server.name}\n"
        f"IP: <code>{server.ip}</code>\n"
        f"Inbound ID: {server.inbound_id}\n"
        f"Лимит: {server.max_users} пользователей",
        reply_markup=keyboard,
        parse_mode="HTML",
    )

    logger.info(f"➕ Админ добавил сервер: {server.name}")


# ============ CALLBACK ОБРАБОТЧИКИ ============


@router.callback_query(F.data.startswith("delete_server:"))
async def delete_server_callback(callback: types.CallbackQuery):
    """Удаление сервера (первый шаг - проверка юзеров)"""
    if callback.from_user.id != settings.ADMIN_USER_ID:
        await callback.answer("❌ Нет доступа", show_alert=True)
        return

    server_id = int(callback.data.split(":")[1])

    async with AsyncSessionLocal() as session:
        server = await get_server_by_id(session, server_id)

        if not server:
            await callback.answer("❌ Сервер не найден", show_alert=True)
            return

        # Проверим кол-во юзеров
        from sqlalchemy import select, func
        from database.models import VpnConfig, User

        result = await session.execute(
            select(func.count(VpnConfig.id)).where(VpnConfig.server_id == server_id)
        )
        user_count = result.scalar()

        if user_count > 0:
            from telegram_bot.keyboards import get_confirm_delete_server_keyboard

            keyboard = get_confirm_delete_server_keyboard(server_id)
            await callback.message.edit_text(
                f"⚠️ На сервере <b>{server.name}</b> есть <b>{user_count}</b> пользователей.\n\n"
                f"Если вы удалите сервер, их текущие конфигурации будут аннулированы, и они автоматически получат новые ключи при следующем запросе через бота.\n\n"
                f"Вы уверены, что хотите удалить этот сервер?",
                reply_markup=keyboard,
                parse_mode="HTML",
            )
        else:
            success = await delete_server(session, server_id)
            if success:
                await callback.message.edit_text(
                    f"✅ Сервер <b>{server.name}</b> удалён", parse_mode="HTML"
                )
                logger.info(f"🗑️ Админ удалил сервер: {server.name}")
            else:
                await callback.answer("❌ Ошибка удаления", show_alert=True)

    await callback.answer()


@router.callback_query(F.data.startswith("confirm_delete_server:"))
async def confirm_delete_server_callback(callback: types.CallbackQuery):
    """Подтверждение удаления сервера с пользователями"""
    if callback.from_user.id != settings.ADMIN_USER_ID:
        await callback.answer("❌ Нет доступа", show_alert=True)
        return

    server_id = int(callback.data.split(":")[1])

    async with AsyncSessionLocal() as session:
        server = await get_server_by_id(session, server_id)
        if not server:
            await callback.message.edit_text("❌ Сервер уже удалён или не найден")
            return

        success = await delete_server(session, server_id, force=True)
        if success:
            await callback.message.edit_text(
                f"✅ Сервер <b>{server.name}</b> и все его конфигурации успешно удалены.",
                parse_mode="HTML",
            )
            logger.info(f"🗑️ Админ принудительно удалил сервер: {server.name}")
        else:
            await callback.answer("❌ Ошибка удаления", show_alert=True)

    await callback.answer()


@router.callback_query(F.data == "cancel_action")
async def cancel_action_callback(callback: types.CallbackQuery):
    """Отмена действия инлайн кнопкой"""
    await callback.message.delete()
    await callback.answer("Отменено")


@router.callback_query(F.data.startswith("export_users:"))
async def export_users_callback(callback: types.CallbackQuery):
    """Выгрузка пользователей сервера в Excel"""
    if callback.from_user.id != settings.ADMIN_USER_ID:
        await callback.answer("❌ Нет доступа", show_alert=True)
        return

    server_id = int(callback.data.split(":")[1])
    await callback.answer("⏳ Создаю Excel файл...", show_alert=False)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        await callback.message.answer(
            "❌ Библиотека openpyxl не установлена на сервере. Обратитесь к разработчику."
        )
        return

    import os
    from datetime import datetime
    from sqlalchemy import select
    from database.models import VpnConfig, User
    from aiogram.types import FSInputFile

    async with AsyncSessionLocal() as session:
        server = await get_server_by_id(session, server_id)
        if not server:
            await callback.answer("❌ Сервер не найден", show_alert=True)
            return

        result = await session.execute(
            select(VpnConfig, User)
            .join(User, VpnConfig.user_id == User.id)
            .where(VpnConfig.server_id == server_id)
        )
        data = result.all()

        if not data:
            await callback.message.answer(
                f"На сервере <b>{server.name}</b> нет пользователей.", parse_mode="HTML"
            )
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Users {server.name}"

        headers = [
            "ID",
            "Username",
            "Имя",
            "Email (X-UI)",
            "VLESS Link",
            "Подписка до",
            "Статус VPN",
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for config, user in data:
            sub_end = (
                user.subscription_end_date.strftime("%Y-%m-%d %H:%M")
                if user.subscription_end_date
                else "Нет"
            )
            status = "Активен" if config.is_active else "Отключен"

            ws.append(
                [
                    user.telegram_id,
                    f"@{user.username}" if user.username else "Нет",
                    user.first_name or "Нет",
                    config.email,
                    config.config_url,
                    sub_end,
                    status,
                ]
            )

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        filename = f"users_{server.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(filename)

        document = FSInputFile(filename)
        await callback.message.answer_document(
            document=document,
            caption=f"📊 Выгрузка пользователей сервера <b>{server.name}</b>",
            parse_mode="HTML",
        )

        try:
            os.remove(filename)
        except OSError:
            pass


@router.callback_query(F.data.startswith("sync_server:"))
async def sync_server_callback(callback: types.CallbackQuery):
    """Синхронизация счётчика сервера"""
    if callback.from_user.id != settings.ADMIN_USER_ID:
        await callback.answer("❌ Нет доступа", show_alert=True)
        return

    server_id = int(callback.data.split(":")[1])

    async with AsyncSessionLocal() as session:
        await sync_server_count(session, server_id)

        server = await get_server_by_id(session, server_id)

        if server:
            await callback.answer(
                f"✅ Синхронизировано: {server.current_users} пользователей",
                show_alert=True,
            )

    await callback.answer()


@router.message(F.text == "🔄 Синхронизировать счётчики")
async def sync_all_servers(message: types.Message):
    """Синхронизировать все серверы"""
    if not is_admin(message):
        return

    async with AsyncSessionLocal() as session:
        servers = await get_all_servers(session)

        for server in servers:
            await sync_server_count(session, server.id)

        await message.answer(
            f"✅ Синхронизировано {len(servers)} серверов", parse_mode="HTML"
        )


@router.message(F.text == "🔙 Назад в админ панель")
async def back_to_admin_panel(message: types.Message):
    """Вернуться в админ панель"""
    if not is_admin(message):
        return

    keyboard = get_admin_panel_keyboard()
    await message.answer(
        "⚙️ <b>Админ панель</b>", reply_markup=keyboard, parse_mode="HTML"
    )


@router.callback_query(F.data.startswith("edit_server:"))
async def edit_server_callback(callback: types.CallbackQuery, state: FSMContext):
    """Начать редактирование сервера"""
    if callback.from_user.id != settings.ADMIN_USER_ID:
        await callback.answer("❌ Нет доступа", show_alert=True)
        return

    server_id = int(callback.data.split(":")[1])

    async with AsyncSessionLocal() as session:
        server = await get_server_by_id(session, server_id)

        if not server:
            await callback.answer("❌ Сервер не найден", show_alert=True)
            return

        # Сохраняем ID сервера в состояние
        await state.update_data(editing_server_id=server_id)

        # Показываем меню редактирования
        keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="📝 Название", callback_data=f"edit_field:name:{server_id}"
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="🌐 IP адрес", callback_data=f"edit_field:ip:{server_id}"
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="🔗 Host", callback_data=f"edit_field:host:{server_id}"
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="👤 Username",
                        callback_data=f"edit_field:username:{server_id}",
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="🔑 Password",
                        callback_data=f"edit_field:password:{server_id}",
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="🔢 Inbound ID",
                        callback_data=f"edit_field:inbound_id:{server_id}",
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="👥 Макс. юзеров",
                        callback_data=f"edit_field:max_users:{server_id}",
                    )
                ],
                # [InlineKeyboardButton(text="⭐ Приоритет", callback_data=f"edit_field:priority:{server_id}")],
                [
                    InlineKeyboardButton(
                        text="🟢 Включить" if not server.is_active else "🔴 Отключить",
                        callback_data=f"toggle_server:{server_id}",
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="❌ Отмена", callback_data=f"cancel_edit:{server_id}"
                    )
                ],
            ]
        )

        status = "🟢 Активен" if server.is_active else "🔴 Отключён"

        await callback.message.edit_text(
            f"<b>⚙️ Редактирование сервера</b>\n\n"
            f"<b>Название:</b> {server.name}\n"
            f"<b>IP:</b> <code>{server.ip}</code>\n"
            f"<b>Host:</b> <code>{server.host}</code>\n"
            f"<b>Username:</b> <code>{server.username}</code>\n"
            f"<b>Password:</b> <code>{'*' * len(server.password)}</code>\n"
            f"<b>Inbound ID:</b> {server.inbound_id}\n"
            f"<b>Макс. юзеров:</b> {server.max_users}\n"
            # f"<b>Приоритет:</b> {server.priority}\n"
            f"<b>Статус:</b> {status}\n\n" f"Выбери что хочешь изменить:",
            reply_markup=keyboard,
            parse_mode="HTML",
        )

    await callback.answer()


@router.callback_query(F.data.startswith("edit_field:"))
async def edit_field_callback(callback: types.CallbackQuery, state: FSMContext):
    """Выбор поля для редактирования"""
    if callback.from_user.id != settings.ADMIN_USER_ID:
        await callback.answer("❌ Нет доступа", show_alert=True)
        return

    _, field, server_id = callback.data.split(":")
    server_id = int(server_id)

    # Сохраняем что редактируем
    await state.update_data(editing_server_id=server_id, editing_field=field)
    await state.set_state(ServerStates.editing_field)

    field_names = {
        "name": "Название сервера",
        "ip": "IP адрес",
        "host": "Host (URL панели)",
        "username": "Username",
        "password": "Password",
        "inbound_id": "Inbound ID",
        "max_users": "Максимум пользователей",
        # "priority": "Приоритет"
    }

    await callback.message.answer(
        f"📝 <b>Редактирование: {field_names[field]}</b>\n\n" f"Введи новое значение:",
        parse_mode="HTML",
        reply_markup=get_cancel_keyboard(),
    )

    await callback.answer()


@router.message(ServerStates.editing_field)
async def process_edit_field(message: types.Message, state: FSMContext):
    """Обработка нового значения поля"""
    if message.text == "❌ Отмена":
        await state.clear()
        keyboard = get_admin_panel_keyboard()
        await message.answer("❌ Редактирование отменено", reply_markup=keyboard)
        return

    data = await state.get_data()
    server_id = data["editing_server_id"]
    field = data["editing_field"]
    new_value = message.text

    async with AsyncSessionLocal() as session:
        # Преобразуем значение если нужно
        if field in ["inbound_id", "max_users"]:
            try:
                new_value = int(new_value)
            except ValueError:
                await message.answer("❌ Значение должно быть числом. Попробуй снова:")
                return

        # Обновляем поле
        update_data = {field: new_value}
        success = await update_server(session, server_id, **update_data)

        if success:
            server = await get_server_by_id(session, server_id)

            field_names = {
                "name": "Название",
                "ip": "IP",
                "host": "Host",
                "username": "Username",
                "password": "Password",
                "inbound_id": "Inbound ID",
                "max_users": "Макс. юзеров",
                # "priority": "Приоритет"
            }

            await message.answer(
                f"✅ <b>{field_names[field]} обновлено!</b>\n\n"
                f"Сервер: <b>{server.name}</b>\n"
                f"Новое значение: <code>{new_value}</code>",
                reply_markup=get_admin_panel_keyboard(),
                parse_mode="HTML",
            )

            logger.info(f"✏️ Админ обновил {field} сервера #{server_id} → {new_value}")
        else:
            await message.answer(
                "❌ Не удалось обновить значение",
                reply_markup=get_admin_panel_keyboard(),
            )

    await state.clear()


@router.callback_query(F.data.startswith("toggle_server:"))
async def toggle_server_callback(callback: types.CallbackQuery):
    """Включить/выключить сервер"""
    if callback.from_user.id != settings.ADMIN_USER_ID:
        await callback.answer("❌ Нет доступа", show_alert=True)
        return

    server_id = int(callback.data.split(":")[1])

    async with AsyncSessionLocal() as session:
        server = await get_server_by_id(session, server_id)

        if not server:
            await callback.answer("❌ Сервер не найден", show_alert=True)
            return

        # Переключаем статус
        new_status = not server.is_active
        await update_server(session, server_id, is_active=new_status)

        status_text = "включён" if new_status else "отключён"
        status_emoji = "🟢" if new_status else "🔴"

        await callback.answer(
            f"✅ Сервер {status_emoji} {status_text}", show_alert=True
        )

        # Обновляем сообщение
        server = await get_server_by_id(session, server_id)

        keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="📝 Название", callback_data=f"edit_field:name:{server_id}"
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="🌐 IP адрес", callback_data=f"edit_field:ip:{server_id}"
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="🔗 Host", callback_data=f"edit_field:host:{server_id}"
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="👤 Username",
                        callback_data=f"edit_field:username:{server_id}",
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="🔑 Password",
                        callback_data=f"edit_field:password:{server_id}",
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="🔢 Inbound ID",
                        callback_data=f"edit_field:inbound_id:{server_id}",
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="👥 Макс. юзеров",
                        callback_data=f"edit_field:max_users:{server_id}",
                    )
                ],
                # [InlineKeyboardButton(text="⭐ Приоритет", callback_data=f"edit_field:priority:{server_id}")],
                [
                    InlineKeyboardButton(
                        text="🟢 Включить" if not server.is_active else "🔴 Отключить",
                        callback_data=f"toggle_server:{server_id}",
                    )
                ],
                [
                    InlineKeyboardButton(
                        text="❌ Отмена", callback_data=f"cancel_edit:{server_id}"
                    )
                ],
            ]
        )

        status = "🟢 Активен" if server.is_active else "🔴 Отключён"

        await callback.message.edit_text(
            f"<b>⚙️ Редактирование сервера</b>\n\n"
            f"<b>Название:</b> {server.name}\n"
            f"<b>IP:</b> <code>{server.ip}</code>\n"
            f"<b>Host:</b> <code>{server.host}</code>\n"
            f"<b>Username:</b> <code>{server.username}</code>\n"
            f"<b>Password:</b> <code>{'*' * len(server.password)}</code>\n"
            f"<b>Inbound ID:</b> {server.inbound_id}\n"
            f"<b>Макс. юзеров:</b> {server.max_users}\n"
            # f"<b>Приоритет:</b> {server.priority}\n"
            f"<b>Статус:</b> {status}\n\n" f"Выбери что хочешь изменить:",
            reply_markup=keyboard,
            parse_mode="HTML",
        )


@router.callback_query(F.data.startswith("cancel_edit:"))
async def cancel_edit_callback(callback: types.CallbackQuery, state: FSMContext):
    """Отмена редактирования"""
    await state.clear()
    await callback.message.delete()
    await callback.answer("Редактирование отменено")


# ============ УДАЛЕНИЕ ПОЛЬЗОВАТЕЛЯ ============


@router.message(F.text == "🗑️ Удалить пользователя")
async def delete_user_start(message: types.Message, state: FSMContext):
    """Начать удаление пользователя"""
    if not is_admin(message):
        return

    await state.set_state(UserManagementStates.waiting_for_delete_username)

    cancel_kb = get_cancel_keyboard()
    await message.answer(
        "🗑️ <b>Удаление пользователя</b>\n\n"
        "Введи username или Telegram ID пользователя:",
        reply_markup=cancel_kb,
        parse_mode="HTML",
    )


@router.message(UserManagementStates.waiting_for_delete_username)
async def process_delete_user(message: types.Message, state: FSMContext):
    """Обработка удаления пользователя"""
    if message.text == "❌ Отмена":
        await state.clear()
        keyboard = get_admin_panel_keyboard()
        await message.answer("❌ Удаление отменено", reply_markup=keyboard)
        return

    search_query = message.text.strip().replace("@", "")

    async with AsyncSessionLocal() as session:
        # Пробуем найти по username
        user = await get_user_by_username(session, search_query)

        # Если не найдено, пробуем по ID
        if not user:
            try:
                telegram_id = int(search_query)
                user = await get_user_by_telegram_id(session, telegram_id)
            except ValueError:
                pass

        if not user:
            await message.answer(
                f"❌ Пользователь <b>{search_query}</b> не найден",
                parse_mode="HTML",
                reply_markup=get_admin_panel_keyboard(),
            )
            await state.clear()
            return

        # Получаем конфиг
        config = await get_user_config(session, user.id)

        server_info = ""

        if config and config.server_id:
            from database.server_operations import get_server_by_id

            server = await get_server_by_id(session, config.server_id)
            if server:
                server_info = f"\n<b>Сервер:</b> {server.name}"

        # Подтверждение удаления
        keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="✅ Да, удалить",
                        callback_data=f"confirm_delete_user:{user.id}",
                    ),
                    InlineKeyboardButton(
                        text="❌ Отмена", callback_data="cancel_delete_user"
                    ),
                ]
            ]
        )

        config_status = "✅ Есть конфиг" if config else "❌ Нет конфига"

        await message.answer(
            f"⚠️ <b>Подтверждение удаления</b>\n\n"
            f"<b>Username:</b> @{user.username or 'нет'}\n"
            f"<b>Имя:</b> {user.first_name or 'нет'}\n"
            f"<b>Telegram ID:</b> <code>{user.telegram_id}</code>\n"
            f"<b>Конфиг:</b> {config_status}{server_info}\n\n"
            f"⚠️ <b>Будет удалено:</b>\n"
            f"- Пользователь из БД\n"
            f"- Конфиг из БД\n"
            f"- Конфиг из 3x-ui панели\n"
            f"- Подписки на каналы\n\n"
            f"<b>Это действие нельзя отменить!</b>",
            reply_markup=keyboard,
            parse_mode="HTML",
        )

    await state.clear()


@router.callback_query(F.data.startswith("confirm_delete_user:"))
async def confirm_delete_user_callback(callback: types.CallbackQuery):
    """Подтверждение удаления пользователя"""
    if callback.from_user.id != settings.ADMIN_USER_ID:
        await callback.answer("❌ Нет доступа", show_alert=True)
        return

    user_id = int(callback.data.split(":")[1])

    async with AsyncSessionLocal() as session:
        # Получаем пользователя
        result = await session.execute(select(User).where(User.id == user_id))
        user = result.scalars().first()

        if not user:
            await callback.answer("❌ Пользователь не найден", show_alert=True)
            return

        # Получаем конфиг
        config = await get_user_config(session, user.id)

        # Удаляем конфиг из 3x-ui
        if config and config.server_id:
            from database.server_operations import (
                get_server_by_id,
                decrement_server_users,
            )
            from xui_api.manager import ServerManagerFactory

            server = await get_server_by_id(session, config.server_id)

            if server:
                try:
                    # Создаём менеджер
                    manager = ServerManagerFactory.create_manager(
                        server.host,
                        server.username,
                        server.password,
                        server.inbound_id,  # ← ДОБАВЬ
                    )

                    # Используем delete_config (не delete_client!)
                    await manager.delete_config(config.email)

                    logger.info(
                        f"🗑️ Удалён конфиг {config.email} из 3x-ui на сервере {server.name}"
                    )

                    # Уменьшаем счётчик
                    await decrement_server_users(session, server.id)

                except Exception as e:
                    logger.error(f"Ошибка удаления из 3x-ui: {e}")

        # Удаляем конфиг
        if config:
            await session.delete(config)

        # Удаляем подписки
        await session.execute(
            delete(UserChannelSubscription).where(
                UserChannelSubscription.user_id == user_id
            )
        )

        # Удаляем пользователя
        await session.delete(user)
        await session.commit()

        await callback.message.edit_text(
            f"✅ <b>Пользователь успешно удалён</b>\n\n"
            f"Username: @{user.username or 'нет'}\n"
            f"Telegram ID: <code>{user.telegram_id}</code>\n\n"
            f"Все данные удалены из БД и 3x-ui панели.",
            parse_mode="HTML",
        )

        await callback.message.answer(
            "👇 Выбери действие:", reply_markup=get_admin_panel_keyboard()
        )

        logger.info(
            f"🗑️ Админ удалил пользователя: @{user.username} (ID: {user.telegram_id})"
        )

    await callback.answer()


@router.callback_query(F.data == "cancel_delete_user")
async def cancel_delete_user_callback(callback: types.CallbackQuery):
    """Отмена удаления"""
    await callback.message.delete()

    # ← ДОБАВЬ ЭТО: Возвращаем админскую клавиатуру
    await callback.message.answer(
        "❌ Удаление отменено", reply_markup=get_admin_panel_keyboard()
    )

    await callback.answer()


@router.message(F.text == "🔔 Уведомить о новых каналах")
async def notify_new_channels_text(message: types.Message):
    """Уведомить всех пользователей о новых каналах (текстовая кнопка)"""
    if message.from_user.id != settings.ADMIN_USER_ID:
        return

    try:
        async with AsyncSessionLocal() as session:
            # Получаем каналы, о которых ещё не уведомили
            result = await session.execute(
                select(Channel)
                .where(Channel.is_active == True)
                .where(Channel.notified_users == False)
                .order_by(Channel.order)
            )
            new_channels = result.scalars().all()

            if not new_channels:
                await message.answer(
                    "ℹ️ <b>Нет новых каналов для уведомления</b>", parse_mode="HTML"
                )
                return

            # Получаем количество пользователей с активными конфигами
            from sqlalchemy import func

            result = await session.execute(
                select(func.count(User.id))
                .join(VpnConfig, User.id == VpnConfig.user_id)
                .where(VpnConfig.is_active == True)
            )
            users_count = result.scalar()

            if users_count == 0:
                await message.answer(
                    "ℹ️ <b>Нет пользователей с активными конфигами</b>",
                    parse_mode="HTML",
                )
                return

            # Формируем список каналов
            channels_list = "\n".join([f"- {ch.display_name}" for ch in new_channels])

            # Показываем подтверждение
            confirm_keyboard = InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text="✅ Да, уведомить", callback_data="confirm_notify_all"
                        )
                    ],
                    [
                        InlineKeyboardButton(
                            text="❌ Отмена", callback_data="cancel_notify_all"
                        )
                    ],
                ]
            )

            await message.answer(
                f"🔔 <b>Уведомление о новых каналах</b>\n\n"
                f"<b>Новых каналов:</b> {len(new_channels)}\n\n"
                f"{channels_list}\n\n"
                f"<b>Будет уведомлено пользователей:</b> {users_count}\n\n"
                f"⚠️ Всем пользователям будут назначены эти каналы с дедлайном подписки.\n\n"
                f"Продолжить?",
                reply_markup=confirm_keyboard,
                parse_mode="HTML",
            )

    except Exception as e:
        logger.error(f"Ошибка при подготовке уведомления: {e}", exc_info=True)
        await message.answer("❌ Ошибка при подготовке уведомления")


@router.callback_query(F.data == "cancel_notify_all")
async def cancel_notify_all_callback(callback: types.CallbackQuery):
    """Отмена уведомления"""
    if callback.from_user.id != settings.ADMIN_USER_ID:
        await callback.answer("❌ Нет доступа", show_alert=True)
        return

    await callback.message.edit_text(
        "❌ <b>Уведомление отменено</b>", parse_mode="HTML"
    )
    await callback.answer()


@router.callback_query(F.data == "confirm_notify_all")
async def confirm_notify_all_callback(callback: types.CallbackQuery):
    """Подтверждение уведомления всех пользователей о новых каналах"""
    if callback.from_user.id != settings.ADMIN_USER_ID:
        await callback.answer("❌ Нет доступа", show_alert=True)
        return

    try:
        # Показываем процесс
        await callback.message.edit_text(
            "⏳ <b>Отправляю уведомления...</b>\n\n"
            "Это может занять некоторое время.",
            parse_mode="HTML",
        )

        async with AsyncSessionLocal() as session:
            # Получаем новые каналы
            result = await session.execute(
                select(Channel)
                .where(Channel.is_active == True)
                .where(Channel.notified_users == False)
                .order_by(Channel.order)
            )
            new_channels = result.scalars().all()

            if not new_channels:
                await callback.message.edit_text(
                    "ℹ️ <b>Нет новых каналов для уведомления</b>", parse_mode="HTML"
                )
                await callback.answer()
                return

            channel_ids = [ch.id for ch in new_channels]

            # Отправляем уведомления
            from database.operations import notify_users_about_new_channels

            await notify_users_about_new_channels(session, callback.bot, channel_ids)

            # Помечаем каналы как отправленные
            for channel in new_channels:
                channel.notified_users = True
            await session.commit()

            # Получаем количество уведомлённых пользователей
            from sqlalchemy import func

            result = await session.execute(
                select(func.count(User.id))
                .join(VpnConfig, User.id == VpnConfig.user_id)
                .where(VpnConfig.is_active == True)
            )
            users_count = result.scalar()

            # Список каналов
            channels_list = "\n".join([f"- {ch.display_name}" for ch in new_channels])

            # ===== ИСПРАВЛЕНИЕ: УБРАЛИ reply_markup =====
            await callback.message.edit_text(
                f"✅ <b>Уведомления отправлены!</b>\n\n"
                f"<b>Каналов:</b> {len(new_channels)}\n\n"
                f"{channels_list}\n\n"
                f"<b>Уведомлено пользователей:</b> {users_count}\n\n"
                f"Пользователи получили уведомление и дедлайн на подписку.",
                parse_mode="HTML",
                # reply_markup удалён, т.к. edit_text не поддерживает ReplyKeyboardMarkup
            )

            logger.info(
                f"✅ Админ уведомил {users_count} пользователей о {len(new_channels)} новых каналах"
            )

    except Exception as e:
        logger.error(f"Ошибка при уведомлении о каналах: {e}")
        await callback.message.edit_text(
            "❌ <b>Произошла ошибка при отправке уведомлений</b>", parse_mode="HTML"
        )
        await callback.answer()


# ============ НАСТРОЙКИ СТОИМОСТИ ============


@router.message(F.text == "💰 Настройки цен")
async def settings_menu(message: types.Message):
    if not is_admin(message):
        return

    async with AsyncSessionLocal() as session:
        price_1m = await get_setting(session, "price_1_month_rub", 150)
        price_6m = await get_setting(session, "price_6_months_rub", 700)

    await message.answer(
        f"💰 <b>Настройки цен тарифов (RUB)</b>\n\n"
        f"🔹 1 месяц: <b>{price_1m}</b>\n"
        f"🔹 6 месяцев: <b>{price_6m}</b>\n\n"
        f"Выберите тариф для изменения:",
        reply_markup=get_settings_keyboard(),
        parse_mode="HTML",
    )


@router.callback_query(F.data.startswith("admin_set_price_"))
async def process_set_price_button(callback: types.CallbackQuery, state: FSMContext):
    if callback.data == "admin_set_price_1m":
        await state.set_state(AdminSettingsStates.waiting_for_price_1m)
        await callback.message.answer(
            "📝 Введите новую цену за 1 месяц (числом):",
            reply_markup=get_cancel_keyboard(),
        )
    elif callback.data == "admin_set_price_6m":
        await state.set_state(AdminSettingsStates.waiting_for_price_6m)
        await callback.message.answer(
            "📝 Введите новую цену за 6 месяцев (числом):",
            reply_markup=get_cancel_keyboard(),
        )
    await callback.answer()


@router.message(AdminSettingsStates.waiting_for_price_1m)
async def process_new_price_1m(message: types.Message, state: FSMContext):
    if message.text == "❌ Отмена":
        await state.clear()
        await message.answer("❌ Отменено", reply_markup=get_admin_panel_keyboard())
        return

    try:
        new_price = int(message.text)
        async with AsyncSessionLocal() as session:
            await set_setting(session, "price_1_month_rub", new_price)

        await state.clear()
        await message.answer(
            f"✅ Цена за 1 месяц успешно изменена на <b>{new_price} руб.</b>",
            reply_markup=get_admin_panel_keyboard(),
            parse_mode="HTML",
        )
    except ValueError:
        await message.answer("❌ Пожалуйста, введите корректное число.")


@router.message(AdminSettingsStates.waiting_for_price_6m)
async def process_new_price_6m(message: types.Message, state: FSMContext):
    if message.text == "❌ Отмена":
        await state.clear()
        await message.answer("❌ Отменено", reply_markup=get_admin_panel_keyboard())
        return

    try:
        new_price = int(message.text)
        async with AsyncSessionLocal() as session:
            await set_setting(session, "price_6_months_rub", new_price)

        await state.clear()
        await message.answer(
            f"✅ Цена за 6 месяцев успешно изменена на <b>{new_price} руб.</b>",
            reply_markup=get_admin_panel_keyboard(),
            parse_mode="HTML",
        )
    except ValueError:
        await message.answer("❌ Пожалуйста, введите корректное число.")


# ============ УПРАВЛЕНИЕ ПОЛЬЗОВАТЕЛЯМИ (ДНИ ПОДПИСКИ) ============


@router.message(F.text == "👥 Управление подписками")
async def user_management_menu(message: types.Message, state: FSMContext):
    if not is_admin(message):
        return
    await state.set_state(UserManagementStates.waiting_for_telegram_id)
    await message.answer(
        "👤 Введите <b>Telegram ID</b> пользователя для управления его подпиской:",
        reply_markup=get_cancel_keyboard(),
        parse_mode="HTML",
    )


@router.message(UserManagementStates.waiting_for_telegram_id)
async def process_user_telegram_id(message: types.Message, state: FSMContext):
    if message.text == "❌ Отмена":
        await state.clear()
        await message.answer("❌ Отменено", reply_markup=get_admin_panel_keyboard())
        return

    try:
        tg_id = int(message.text)
        async with AsyncSessionLocal() as session:
            user = await get_user_by_telegram_id(session, tg_id)
            if not user:
                await message.answer("❌ Пользователь с таким ID не найден.")
                return

            end_date = user.subscription_end_date
            import datetime

            if end_date:
                end_str = end_date.strftime("%d.%m.%Y %H:%M")
                is_active = (
                    "✅ Активна"
                    if end_date > datetime.datetime.utcnow()
                    else "⛔️ Истекла"
                )
            else:
                end_str = "Нет активной подписки"
                is_active = "⛔️ Нет"

            await state.clear()
            await message.answer(
                f"👤 <b>Профиль пользователя {tg_id}</b>\n\n"
                f"Статус: {is_active}\n"
                f"Дата окончания: <b>{end_str}</b>\n\n"
                f"Выберите действие:",
                reply_markup=get_user_manage_keyboard(tg_id),
                parse_mode="HTML",
            )

    except ValueError:
        await message.answer("❌ Введите корректный числовой ID.")


@router.callback_query(F.data.startswith("admin_add_days_"))
async def process_add_days_button(callback: types.CallbackQuery, state: FSMContext):
    tg_id = int(callback.data.split("_")[-1])
    await state.update_data(target_user_id=tg_id)
    await state.set_state(UserManagementStates.waiting_for_add_days)
    await callback.message.answer(
        f"➕ Сколько <b>дней</b> добавить пользователю {tg_id}?",
        reply_markup=get_cancel_keyboard(),
        parse_mode="HTML",
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_remove_days_"))
async def process_remove_days_button(callback: types.CallbackQuery, state: FSMContext):
    tg_id = int(callback.data.split("_")[-1])
    await state.update_data(target_user_id=tg_id)
    await state.set_state(UserManagementStates.waiting_for_remove_days)
    await callback.message.answer(
        f"➖ Сколько <b>дней</b> отнять у пользователя {tg_id}?",
        reply_markup=get_cancel_keyboard(),
        parse_mode="HTML",
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin_disable_sub_"))
async def process_disable_sub_button(callback: types.CallbackQuery, state: FSMContext):
    tg_id = int(callback.data.split("_")[-1])
    async with AsyncSessionLocal() as session:
        user = await get_user_by_telegram_id(session, tg_id)
        if user:
            # Обнуляем подписку
            import datetime

            user.subscription_end_date = (
                datetime.datetime.utcnow() - datetime.timedelta(days=1)
            )
            await session.commit()

            # Отключаем конфиг в XUI
            config = await get_user_config(session, user.id)
            if config:
                if config.server_id:
                    from database.server_operations import get_server_by_id

                    server = await get_server_by_id(session, config.server_id)
                    if server:
                        from xui_api.manager import ServerManagerFactory

                        manager = ServerManagerFactory.create_manager(
                            server.host,
                            server.username,
                            server.password,
                            server.inbound_id,
                        )
                    else:
                        from xui_api.manager import Xui3Manager

                        manager = Xui3Manager()
                else:
                    from xui_api.manager import Xui3Manager

                    manager = Xui3Manager()
                await manager.disable_config(config.email)
                config.is_active = False
                await session.commit()

            await callback.message.answer(
                f"✅ Подписка пользователя {tg_id} успешно выключена, конфиг деактивирован.",
                reply_markup=get_admin_panel_keyboard(),
            )
    await callback.answer()


@router.message(UserManagementStates.waiting_for_add_days)
async def process_add_days_input(message: types.Message, state: FSMContext):
    if message.text == "❌ Отмена":
        await state.clear()
        await message.answer("❌ Отменено", reply_markup=get_admin_panel_keyboard())
        return

    try:
        days = int(message.text)
        data = await state.get_data()
        tg_id = data.get("target_user_id")

        import datetime

        async with AsyncSessionLocal() as session:
            user = await get_user_by_telegram_id(session, tg_id)
            if (
                not user.subscription_end_date
                or user.subscription_end_date < datetime.datetime.utcnow()
            ):
                user.subscription_end_date = (
                    datetime.datetime.utcnow() + datetime.timedelta(days=days)
                )
            else:
                user.subscription_end_date += datetime.timedelta(days=days)
            await session.commit()

        await state.clear()
        await message.answer(
            f"✅ Успешно добавлено {days} дней пользователю {tg_id}.",
            reply_markup=get_admin_panel_keyboard(),
        )

    except ValueError:
        await message.answer("❌ Введите число.")


@router.message(UserManagementStates.waiting_for_remove_days)
async def process_remove_days_input(message: types.Message, state: FSMContext):
    if message.text == "❌ Отмена":
        await state.clear()
        await message.answer("❌ Отменено", reply_markup=get_admin_panel_keyboard())
        return

    try:
        days = int(message.text)
        data = await state.get_data()
        tg_id = data.get("target_user_id")

        import datetime

        async with AsyncSessionLocal() as session:
            user = await get_user_by_telegram_id(session, tg_id)
            if user.subscription_end_date:
                user.subscription_end_date -= datetime.timedelta(days=days)

                # Если после вычитания подписка истекла, рубим конфиг
                if user.subscription_end_date < datetime.datetime.utcnow():
                    config = await get_user_config(session, user.id)
                    if config:
                        if config.server_id:
                            from database.server_operations import get_server_by_id

                            server = await get_server_by_id(session, config.server_id)
                            if server:
                                from xui_api.manager import ServerManagerFactory

                                manager = ServerManagerFactory.create_manager(
                                    server.host,
                                    server.username,
                                    server.password,
                                    server.inbound_id,
                                )
                            else:
                                from xui_api.manager import Xui3Manager

                                manager = Xui3Manager()
                        else:
                            from xui_api.manager import Xui3Manager

                            manager = Xui3Manager()
                        await manager.disable_config(config.email)
                        config.is_active = False

                await session.commit()
                await message.answer(
                    f"✅ Успешно отнято {days} дней у пользователя {tg_id}.",
                    reply_markup=get_admin_panel_keyboard(),
                )
            else:
                await message.answer("❌ У пользователя и так нет подписки.")
        await state.clear()

    except ValueError:
        await message.answer("❌ Введите число.")


# ============ СОЗДАТЬ РЕКЛАМНЫЙ ПОСТ ============


@router.message(F.text == "📣 Создать рекламный пост")
async def create_ad_post(message: types.Message):
    """Генерирует рекламный пост с фото и кнопкой"""
    if not is_admin(message):
        return

    ad_text = (
        "🔥 <b>Надоели медленные VPN и бесконечная реклама?</b>\n\n"
        "Забудьте про дорогие подписки и постоянные ограничения.\n\n"
        "Есть Telegram-бот, который открывает доступ ко всем заблокированным сервисам абсолютно бесплатно.\n\n"
        "⚡️ Скорость как у платного VPN\n"
        "🛡 Надежный протокол VLESS — стабильный обход блокировок\n"
        "📶 Работает 24/7 без тормозов\n"
        "🚫 Без рекламы и скрытых платежей\n\n"
        "Просто нажмите кнопку ниже и получите доступ к свободному интернету за 1 минуту 👇"
    )

    import os
    from aiogram.types import FSInputFile, InlineKeyboardMarkup, InlineKeyboardButton

    image_path = os.path.join("images", "ad.png")

    # Создаем инлайн клавиатуру с кнопкой-ссылкой на бота
    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="🚀 Получить бесплатный VPN",
                    url="https://t.me/free_v1ess_vpn_bot",
                )
            ]
        ]
    )

    if os.path.exists(image_path):
        photo = FSInputFile(image_path)
        await message.answer_photo(
            photo=photo, caption=ad_text, parse_mode="HTML", reply_markup=keyboard
        )
    else:
        # Если фото нет, отправляем просто текстом
        await message.answer(
            f"<i>⚠️ Внимание админ: Файл {image_path} не найден! Отправляю без фото.</i>\n\n"
            + ad_text,
            parse_mode="HTML",
            reply_markup=keyboard,
        )
